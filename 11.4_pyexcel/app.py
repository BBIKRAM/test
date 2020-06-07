import openpyxl

# to create workbook
# wb = openpyxl.workbook()          #COMMAND
wb = openpyxl.load_workbook("test.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

# wb.create_sheet("sheet2",0)
# wb.remove_sheet(sheet)
# cell = sheet["A1"]

#QUERY
cell = sheet.cell(row=1,column=3)
print(cell.value)
print(cell.row)
print(cell.column)
print(cell.coordinate)


# print(sheet.max_column)
# print(sheet.max_row)


# for row in range(1 , sheet.max_row + 1):
#     for column in range(1 , sheet.max_column + 1):
#         cell = sheet.cell(row , column)
#         print(cell.value)
        

# cell = sheet["a1"]
# print(cell)
# column = sheet["a"]
# print(column)
# cells = sheet["a:c"]
# cells = sheet["a1:c3"]
# sheet[1:3]
# print(cells)


sheet.append([1,2,3])
wb.save("test2.xlsx")
# sheet.insert_cols
# sheet.delete_cols


#COMMADN QUERY SEPERTAION METHOD
#COMMAND : make change
#query: read only no change